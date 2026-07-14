"""
Report generation endpoints.

POST /reports/generate             — trigger async report generation
GET  /reports/                     — list user's generated reports
GET  /reports/{report_id}          — get report status/metadata
GET  /reports/{report_id}/download — stream report file download
POST /reports/export/blades        — synchronous Excel export for small blade sets
"""

from __future__ import annotations

import io
import os
import uuid
from typing import Annotated, Any

import structlog
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_current_user
from app.db.session import get_db
from app.models.enums import ReportStatus, ReportType
from app.schemas.base import PaginatedResponse
from app.schemas.report import ReportGenerateRequest, ReportResponse

logger = structlog.get_logger(__name__)
router = APIRouter()

# MIME types per report format
_MIME_TYPES: dict[ReportType, str] = {
    ReportType.PDF: "application/pdf",
    ReportType.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

_FILE_EXTENSIONS: dict[ReportType, str] = {
    ReportType.PDF: ".pdf",
    ReportType.EXCEL: ".xlsx",
}


# ---------------------------------------------------------------------------
# Background task: generate report
# ---------------------------------------------------------------------------


async def _generate_report_task(report_id: uuid.UUID) -> None:
    """
    Background coroutine: generates a real PDF or Excel report file and
    updates the Report record with status=READY (or FAILED on error).
    """
    import os
    from datetime import datetime, timezone
    from pathlib import Path

    from app.db.session import AsyncSessionLocal
    from app.models.report import Report
    from app.reports.generator import ReportGenerator

    async with AsyncSessionLocal() as db:
        try:
            result = await db.execute(select(Report).where(Report.id == report_id))
            report = result.scalar_one_or_none()
            if report is None:
                logger.error("report_not_found_in_bg_task", report_id=str(report_id))
                return

            report.status = ReportStatus.GENERATING
            await db.commit()

            generator = ReportGenerator()
            filter_params: dict = report.filter_params or {}

            report_type = ReportType(report.report_type)
            if report_type == ReportType.EXCEL:
                file_bytes = await generator.generate_blade_report_excel([], db, filter_params)
                ext = ".xlsx"
            else:
                file_bytes = await generator.generate_blade_report_pdf([], db, filter_params)
                ext = ".pdf"

            reports_dir = Path(os.environ.get("REPORTS_DIR", "/app/reports"))
            reports_dir.mkdir(parents=True, exist_ok=True)
            file_path = reports_dir / f"report_{report_id}{ext}"
            file_path.write_bytes(file_bytes)

            report.status = ReportStatus.READY
            report.file_path = str(file_path)
            report.file_size_bytes = len(file_bytes)
            report.completed_at = datetime.now(timezone.utc)
            await db.commit()

            logger.info("report_generated", report_id=str(report_id), size_bytes=len(file_bytes))

        except Exception as exc:  # noqa: BLE001
            logger.error("report_generation_failed", report_id=str(report_id), error=str(exc))
            try:
                result = await db.execute(select(Report).where(Report.id == report_id))
                report = result.scalar_one_or_none()
                if report:
                    report.status = ReportStatus.FAILED
                    report.error_message = str(exc)
                    await db.commit()
            except Exception:  # noqa: BLE001
                pass


# ---------------------------------------------------------------------------
# POST /generate
# ---------------------------------------------------------------------------


@router.post(
    "/generate",
    response_model=ReportResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Trigger asynchronous report generation",
)
async def generate_report(
    body: ReportGenerateRequest,
    background_tasks: BackgroundTasks,
    current_user: Annotated[Any, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> Any:
    """
    Queue a new report for generation.

    Returns immediately with ``status=PENDING``.  Poll
    ``GET /reports/{report_id}`` until ``status`` becomes ``READY`` or
    ``FAILED``.

    The report is generated asynchronously in a background task.
    For production workloads, replace the background task with a Celery
    task to avoid blocking the event loop.
    """
    from app.models.report import Report

    report = Report(
        name=body.name,
        report_type=body.report_type,
        status=ReportStatus.PENDING,
        generated_by_id=current_user.id,
        filter_params=body.filters.model_dump(mode="json"),
    )
    db.add(report)
    await db.commit()
    await db.refresh(report)

    background_tasks.add_task(_generate_report_task, report.id)

    logger.info(
        "report_queued",
        report_id=str(report.id),
        name=body.name,
        type=str(body.report_type),
        user_id=str(current_user.id),
    )
    return report


# ---------------------------------------------------------------------------
# GET /
# ---------------------------------------------------------------------------


@router.get(
    "/",
    response_model=PaginatedResponse[ReportResponse],
    status_code=status.HTTP_200_OK,
    summary="List reports generated by the current user",
)
async def list_reports(
    current_user: Annotated[Any, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    report_status: ReportStatus | None = Query(default=None, alias="status"),
) -> Any:
    """
    Return a paginated list of reports created by the current user,
    ordered by creation date (newest first).

    Optionally filter by ``status``: PENDING, GENERATING, READY, or FAILED.
    """
    from app.models.report import Report

    conditions = [Report.generated_by_id == current_user.id]
    if report_status is not None:
        conditions.append(Report.status == report_status)

    total: int = (
        await db.execute(
            select(func.count()).select_from(Report).where(*conditions)
        )
    ).scalar_one()

    items = list(
        (
            await db.execute(
                select(Report)
                .where(*conditions)
                .order_by(Report.created_at.desc())
                .offset(skip)
                .limit(limit)
            )
        )
        .scalars()
        .all()
    )

    page = skip // limit + 1 if limit > 0 else 1
    return PaginatedResponse(items=items, total=total, page=page, page_size=limit)


# ---------------------------------------------------------------------------
# GET /{report_id}
# ---------------------------------------------------------------------------


@router.get(
    "/{report_id}",
    response_model=ReportResponse,
    status_code=status.HTTP_200_OK,
    summary="Get report status and metadata",
)
async def get_report(
    report_id: uuid.UUID,
    current_user: Annotated[Any, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> Any:
    """
    Return the metadata record for a report, including its generation status.

    Poll this endpoint after calling ``POST /reports/generate`` to check
    when ``status`` transitions to ``READY``.

    Raises:
        HTTP 404 — report not found or does not belong to the current user.
    """
    from app.models.report import Report

    result = await db.execute(
        select(Report).where(
            Report.id == report_id,
            Report.generated_by_id == current_user.id,
        )
    )
    report = result.scalar_one_or_none()
    if report is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found",
        )

    return report


# ---------------------------------------------------------------------------
# DELETE /{report_id}
# ---------------------------------------------------------------------------


@router.delete(
    "/{report_id}",
    status_code=status.HTTP_200_OK,
    summary="Delete a report record and its file",
)
async def delete_report(
    report_id: uuid.UUID,
    current_user: Annotated[Any, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    """Delete a report record. Also removes the file from disk if it exists."""
    import os
    from app.models.report import Report

    result = await db.execute(
        select(Report).where(
            Report.id == report_id,
            Report.generated_by_id == current_user.id,
        )
    )
    report = result.scalar_one_or_none()
    if report is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found",
        )

    # Remove file from disk if it exists
    if report.file_path:
        try:
            if os.path.exists(report.file_path):
                os.remove(report.file_path)
        except OSError:
            pass  # Non-fatal — still delete the DB record

    await db.delete(report)
    await db.commit()

    logger.info("report_deleted", report_id=str(report_id), by=str(current_user.id))
    return {"success": True, "message": f"Report {report_id} deleted"}


# ---------------------------------------------------------------------------
# GET /{report_id}/download
# ---------------------------------------------------------------------------


@router.get(
    "/{report_id}/download",
    status_code=status.HTTP_200_OK,
    summary="Stream-download a completed report file",
    responses={
        200: {
            "content": {
                "application/pdf": {},
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {},
            },
            "description": "The generated report file",
        }
    },
)
async def download_report(
    report_id: uuid.UUID,
    current_user: Annotated[Any, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> StreamingResponse:
    """
    Stream the generated report file to the client.

    Only available when ``status=READY``.  The ``Content-Disposition``
    header is set to ``attachment`` with the original report name.

    Raises:
        HTTP 404 — report not found.
        HTTP 409 — report is not yet ready (still pending/generating).
        HTTP 502 — report file not found on storage.
    """
    from app.models.report import Report

    result = await db.execute(
        select(Report).where(
            Report.id == report_id,
            Report.generated_by_id == current_user.id,
        )
    )
    report = result.scalar_one_or_none()
    if report is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Report {report_id} not found",
        )

    if report.status != ReportStatus.READY:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Report is not ready for download (status: {report.status})",
        )

    if not report.file_path:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="Report file path is missing",
        )

    mime_type = _MIME_TYPES.get(ReportType(report.report_type), "application/octet-stream")
    ext = _FILE_EXTENSIONS.get(ReportType(report.report_type), ".bin")
    safe_name = f"{report.name.replace(' ', '_')}{ext}"

    import aiofiles
    from pathlib import Path

    file_path = Path(report.file_path)
    if not file_path.exists():
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Report file not found on storage: {report.file_path}",
        )

    async def _file_iterator():  # noqa: ANN202
        async with aiofiles.open(file_path, "rb") as f:
            while chunk := await f.read(64 * 1024):
                yield chunk

    return StreamingResponse(
        content=_file_iterator(),
        media_type=mime_type,
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}"',
            "Content-Length": str(report.file_size_bytes or 0),
        },
    )


# ---------------------------------------------------------------------------
# POST /export/blades
# ---------------------------------------------------------------------------


@router.post(
    "/export/blades",
    status_code=status.HTTP_200_OK,
    summary="Synchronous Excel export of the current blade list",
    responses={
        200: {
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}
            },
            "description": "Excel workbook with blade records",
        }
    },
)
async def export_blades_excel(
    current_user: Annotated[Any, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    blade_status: Any | None = Query(default=None, alias="status"),
    limit: int = Query(
        default=1000,
        ge=1,
        le=5000,
        description="Maximum number of blades to export (capped at 5000)",
    ),
) -> StreamingResponse:
    """
    Generate and immediately return an Excel workbook containing blade
    records matching the supplied filters.

    Intended for small-to-medium exports.  For large datasets use the
    async ``POST /reports/generate`` flow instead.

    Requires ``openpyxl`` to be installed.

    Raises:
        HTTP 400 — openpyxl not available.
    """
    try:
        import openpyxl  # type: ignore[import]
        from openpyxl.styles import Font, PatternFill  # type: ignore[import]
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="openpyxl is not installed. Install it to use Excel export.",
        )

    from app.models.blade import Blade
    from app.models.enums import BladeStatus as BS

    conditions = [Blade.deleted_at.is_(None)]
    if blade_status is not None:
        conditions.append(Blade.status == blade_status)

    blades = list(
        (
            await db.execute(
                select(Blade)
                .where(*conditions)
                .order_by(Blade.created_at.desc())
                .limit(limit)
            )
        )
        .scalars()
        .all()
    )

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blades"

    headers = [
        "Serial Number", "Melt Number", "Part Number", "Work Order",
        "Status", "Station", "Engine Number", "Running Hours",
        "Created At", "Updated At",
    ]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_num, blade in enumerate(blades, start=2):
        ws.cell(row=row_num, column=1, value=blade.serial_number)
        ws.cell(row=row_num, column=2, value=blade.melt_number)
        ws.cell(row=row_num, column=3, value=blade.part_number)
        ws.cell(row=row_num, column=4, value=blade.work_order_number)
        ws.cell(row=row_num, column=5, value=str(blade.status))
        ws.cell(
            row=row_num,
            column=6,
            value=blade.current_station.name if blade.current_station else None,
        )
        ws.cell(row=row_num, column=7, value=blade.engine_number)
        ws.cell(row=row_num, column=8, value=float(blade.running_hours) if blade.running_hours else None)
        ws.cell(row=row_num, column=9, value=blade.created_at.isoformat() if blade.created_at else None)
        ws.cell(row=row_num, column=10, value=blade.updated_at.isoformat() if blade.updated_at else None)

    # Auto-width columns
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(
        "blades_exported_excel",
        user_id=str(current_user.id),
        row_count=len(blades),
    )

    return StreamingResponse(
        content=iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="blade_export.xlsx"',
        },
    )


# ---------------------------------------------------------------------------
# POST /export/hptr-slots
# ---------------------------------------------------------------------------


@router.post(
    "/export/hptr-slots",
    status_code=status.HTTP_200_OK,
    summary="Synchronous Excel export of a work order's saved HPTR slot assignments",
    responses={
        200: {
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}
            },
            "description": "Excel workbook with all 90 of the work order's saved HPTR slot assignments in one sheet",
        }
    },
)
async def export_hptr_slots_excel(
    current_user: Annotated[Any, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    work_order_number: str = Query(..., description="Work order number whose saved HPTR slots to export"),
) -> StreamingResponse:
    """
    Generate and immediately return an Excel workbook containing the given
    work order's saved (active) HPTR slot assignments — all rows in one sheet,
    sorted by slot number, with a "Half" column marking W1 vs W2.

    Raises:
        HTTP 400 — openpyxl not available.
        HTTP 404 — no active HPTR slot allocation found for this work order.
    """
    try:
        import openpyxl  # type: ignore[import]
        from openpyxl.styles import Font, PatternFill  # type: ignore[import]
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="openpyxl is not installed. Install it to use Excel export.",
        )

    from app.models.blade import Blade
    from app.models.enums import BladeType
    from app.models.measurement import Measurement
    from app.models.slot_allocation import SlotAllocation

    rows = (
        await db.execute(
            select(Blade, SlotAllocation)
            .join(SlotAllocation, SlotAllocation.blade_id == Blade.id)
            .where(
                Blade.work_order_number == work_order_number,
                Blade.blade_type == BladeType.HPTR,
                Blade.deleted_at.is_(None),
                SlotAllocation.is_active.is_(True),
            )
        )
    ).all()

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No active HPTR slot allocation found for work order '{work_order_number}'",
        )

    blade_ids = [blade.id for blade, _ in rows]
    subq = (
        select(
            Measurement.blade_id,
            func.max(Measurement.measured_at).label("latest_at"),
        )
        .where(Measurement.blade_id.in_(blade_ids), Measurement.measurement_type == "INITIAL")
        .group_by(Measurement.blade_id)
        .subquery()
    )
    meas_rows = (
        await db.execute(
            select(Measurement.blade_id, Measurement.weight_grams, Measurement.static_moment_gcm)
            .join(
                subq,
                (Measurement.blade_id == subq.c.blade_id)
                & (Measurement.measured_at == subq.c.latest_at),
            )
        )
    ).all()
    meas_map = {r.blade_id: r for r in meas_rows}

    # Matches the OH Slot Allocation page's default 90-slot rotor split
    # (45/45) — the operator-entered "Total Slots on Rotor" value isn't
    # persisted per-batch, so this mirrors what the UI itself falls back to
    # once a saved batch is reopened in a fresh session.
    half = 45

    def _slot_num(pair: tuple) -> int:
        s = pair[1].slot_number
        return int(s) if s.isdigit() else 0

    def _weight(pair: tuple) -> float:
        meas = meas_map.get(pair[0].id)
        return float(meas.weight_grams) if meas and meas.weight_grams is not None else 0.0

    w1_rows = sorted((r for r in rows if _slot_num(r) <= half), key=_slot_num)
    w2_rows = sorted((r for r in rows if _slot_num(r) > half), key=_slot_num)

    w1_total = sum(_weight(r) for r in w1_rows)
    w2_total = sum(_weight(r) for r in w2_rows)
    x_diff = abs(w1_total - w2_total)

    # y (the rotor's own pre-existing unbalance, entered as "Rotor Unbalance
    # (g)" when slots were saved) isn't persisted on SlotAllocation/WorkOrder
    # — recover it from the WorkflowLog remark written at save time.
    import re

    from app.models.enums import BladeStatus as BS
    from app.models.workflow import WorkflowLog

    log_remarks = (
        await db.execute(
            select(WorkflowLog.remarks)
            .join(Blade, Blade.id == WorkflowLog.blade_id)
            .where(
                Blade.work_order_number == work_order_number,
                Blade.blade_type == BladeType.HPTR,
                WorkflowLog.to_status == BS.SLOT_ASSIGNED,
            )
            .order_by(WorkflowLog.timestamp.desc())
            .limit(1)
        )
    ).scalar_one_or_none()

    start_slot_val: int | None = None
    y_unbalance: float | None = None
    if log_remarks:
        m_start = re.search(r"start slot (\d+)", log_remarks)
        if m_start:
            start_slot_val = int(m_start.group(1))
        m_unbal = re.search(r"unbalance ([\d.]+) g", log_remarks)
        if m_unbal:
            y_unbalance = float(m_unbal.group(1))

    y_for_calc = y_unbalance if y_unbalance is not None else 0.0
    adjusted_diff = abs(x_diff - y_for_calc)
    target_min, target_max = 1.5, 2.0
    is_balanced = target_min <= adjusted_diff <= target_max

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    label_font = Font(bold=True)
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HPTR Slot Assignments"

    ws.cell(row=1, column=1, value=f"Work Order {work_order_number} — HPTR Set Making Summary").font = Font(bold=True, size=13)

    summary = [
        ("Start Slot", start_slot_val if start_slot_val is not None else "Not recorded"),
        ("W1 Total (g)", round(w1_total, 2)),
        ("W2 Total (g)", round(w2_total, 2)),
        ("x = |W1 - W2| (g)", round(x_diff, 2)),
        ("y = Rotor Unbalance (g)", round(y_unbalance, 2) if y_unbalance is not None else "Not recorded"),
        ("|x - y| (g)", round(adjusted_diff, 2)),
        ("Target band (g)", f"{target_min}-{target_max}"),
        ("Status", "Balanced" if is_balanced else "Not balanced"),
    ]
    for i, (label, value) in enumerate(summary):
        r = 3 + i
        ws.cell(row=r, column=1, value=label).font = label_font
        cell = ws.cell(row=r, column=2, value=value)
        if label == "Status":
            cell.fill = ok_fill if is_balanced else warn_fill
            cell.font = Font(bold=True)

    table_headers = ["Slot", "Blade Serial", "Melt No.", "Weight (g)", "Static Moment (g-cm)"]
    table_start_row = 3 + len(summary) + 2
    w1_col_start, w2_col_start = 1, len(table_headers) + 2  # one blank column between the two blocks

    def _write_table(col_start: int, title: str, table_rows: list) -> None:
        ws.cell(row=table_start_row, column=col_start, value=title).font = Font(bold=True, size=12)
        header_row = table_start_row + 1
        for i, header in enumerate(table_headers):
            cell = ws.cell(row=header_row, column=col_start + i, value=header)
            cell.fill = header_fill
            cell.font = header_font
        for row_offset, (blade, alloc) in enumerate(table_rows, start=1):
            meas = meas_map.get(blade.id)
            r = header_row + row_offset
            ws.cell(row=r, column=col_start, value=int(alloc.slot_number) if alloc.slot_number.isdigit() else alloc.slot_number)
            ws.cell(row=r, column=col_start + 1, value=blade.serial_number)
            ws.cell(row=r, column=col_start + 2, value=blade.melt_number)
            ws.cell(row=r, column=col_start + 3, value=float(meas.weight_grams) if meas and meas.weight_grams is not None else None)
            ws.cell(row=r, column=col_start + 4, value=float(meas.static_moment_gcm) if meas and meas.static_moment_gcm is not None else None)

    _write_table(w1_col_start, f"W1 (Slots 1-{half})", w1_rows)
    _write_table(w2_col_start, f"W2 (Slots {half + 1}-90)", w2_rows)

    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(
        "hptr_slots_exported_excel",
        user_id=str(current_user.id),
        work_order_number=work_order_number,
        row_count=len(rows),
    )

    return StreamingResponse(
        content=iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="hptr_slots_{work_order_number}.xlsx"',
        },
    )
