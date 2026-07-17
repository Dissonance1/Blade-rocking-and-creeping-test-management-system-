"""
Report generator — Excel (openpyxl) and PDF (ReportLab) outputs.

All public methods are async and return raw bytes so the caller can
stream the response directly without writing to disk.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any
from uuid import UUID

import structlog
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

logger = structlog.get_logger(__name__)

# ---------------------------------------------------------------------------
# Colour palette (openpyxl / ReportLab)
# ---------------------------------------------------------------------------

# Excel cell fill colours (hex without #)
_XL_HEADER_FILL = "1F4E79"       # dark blue header
_XL_SUBHEADER_FILL = "2E75B6"    # medium blue sub-header
_XL_ALT_ROW_FILL = "D6E4F0"      # light blue alternating row

_STATUS_COLOURS: dict[str, str] = {
    "RECEIVED": "BDD7EE",
    "IN_PROGRESS": "FFE699",
    "COMPLETED": "C6EFCE",
    "REJECTED": "FFC7CE",
    "PENDING": "FFEB9C",
}

# ReportLab colours
_RL_HEADER_BG = (0.122, 0.306, 0.475)   # RGB floats 0-1
_RL_ALT_ROW_BG = (0.839, 0.894, 0.941)


class ReportGenerator:
    """Generate Excel and PDF blade reports plus dashboard summary stats."""

    # ==================================================================
    # Excel report
    # ==================================================================

    async def generate_blade_report_excel(
        self,
        blade_ids: list[UUID],
        db: AsyncSession,
        filter_params: dict | None = None,
    ) -> bytes:
        """
        Build a styled multi-sheet Excel workbook for *blade_ids*.

        Sheets
        ------
        * **Summary** — one row per blade with key metadata
        * **Measurements** — all measurement records for included blades
        * **Slot Allocations** — slot assignment history
        * **Workflow History** — ordered workflow state transitions

        Returns
        -------
        bytes
            Raw ``.xlsx`` bytes, ready to stream as a download.
        """
        try:
            import openpyxl  # type: ignore[import]
            from openpyxl.styles import (
                Alignment,
                Border,
                Font,
                PatternFill,
                Side,
            )
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for Excel reports.") from exc

        blades, measurements, slots, workflow_events = await self._fetch_report_data(
            blade_ids, db, filter_params
        )

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        # ---- helper closures ----

        def header_font() -> Font:
            return Font(bold=True, color="FFFFFF", size=11)

        def header_fill(hex_colour: str) -> PatternFill:
            return PatternFill("solid", fgColor=hex_colour)

        def thin_border() -> Border:
            thin = Side(style="thin", color="CCCCCC")
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        def alt_fill(row_idx: int) -> PatternFill | None:
            if row_idx % 2 == 0:
                return PatternFill("solid", fgColor=_XL_ALT_ROW_FILL)
            return None

        def write_headers(ws: Any, headers: list[str]) -> None:
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font()
                cell.fill = header_fill(_XL_HEADER_FILL)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border()

        def auto_width(ws: Any) -> None:
            for column_cells in ws.columns:
                max_len = max(
                    (len(str(cell.value)) if cell.value else 0) for cell in column_cells
                )
                col_letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        # ---- Sheet 1: Summary ----
        ws_summary = wb.create_sheet("Summary")
        ws_summary.row_dimensions[1].height = 22
        summary_headers = [
            "Serial Number",
            "Melt Number",
            "Status",
            "Station",
            "Created At",
            "Last Updated",
            "Assigned To",
        ]
        write_headers(ws_summary, summary_headers)
        for row_idx, blade in enumerate(blades, start=2):
            row_data = [
                getattr(blade, "serial_number", ""),
                getattr(blade, "melt_number", ""),
                _ev(getattr(blade, "status", "")),
                str(getattr(blade, "current_station_id", "") or ""),
                _fmt_dt(getattr(blade, "created_at", None)),
                _fmt_dt(getattr(blade, "updated_at", None)),
                "",
            ]
            fill = _STATUS_COLOURS.get(str(getattr(blade, "status", "")), None)
            for col, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border()
                if col == 3 and fill:  # colour the status cell
                    cell.fill = PatternFill("solid", fgColor=fill)
                elif alt_fill(row_idx):
                    cell.fill = alt_fill(row_idx)
        auto_width(ws_summary)

        # blade_id → blade lookup for serial/melt lookups in sub-sheets
        blade_by_id_xl = {str(getattr(b, "id", "")): b for b in blades}

        # ---- Sheet 2: Measurements ----
        ws_meas = wb.create_sheet("Measurements")
        meas_headers = [
            "Serial Number", "Melt Number", "Measurement Type",
            "Weight (g)", "Static Moment (g·cm)",
            "Rocking Value", "Creep Value", "Measured At", "Notes",
        ]
        write_headers(ws_meas, meas_headers)
        for row_idx, meas in enumerate(measurements, start=2):
            blade = blade_by_id_xl.get(str(getattr(meas, "blade_id", "")))
            w  = getattr(meas, "weight_grams", None)
            sm = getattr(meas, "static_moment_gcm", None)
            rv = getattr(meas, "rocking_value", None)
            cv = getattr(meas, "creep_value", None)
            row_data = [
                getattr(blade, "serial_number", "") if blade else "",
                getattr(blade, "melt_number",   "") if blade else "",
                _ev(getattr(meas, "measurement_type", "")),
                round(float(w),  2) if w  is not None else "",
                round(float(sm), 2) if sm is not None else "",
                round(float(rv), 4) if rv is not None else "",
                round(float(cv), 4) if cv is not None else "",
                _fmt_dt(getattr(meas, "measured_at", None)),
                getattr(meas, "notes", "") or "",
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws_meas.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border()
                if alt_fill(row_idx):
                    cell.fill = alt_fill(row_idx)
        auto_width(ws_meas)

        # ---- Sheet 3: Slot Allocations ----
        ws_slots = wb.create_sheet("Slot Allocations")
        slot_headers = [
            "Serial Number", "Slot Number", "Position",
            "Balanced", "Imbalance Value", "Remarks", "Allocated At",
        ]
        write_headers(ws_slots, slot_headers)
        for row_idx, slot in enumerate(slots, start=2):
            blade = blade_by_id_xl.get(str(getattr(slot, "blade_id", "")))
            unbal = getattr(slot, "unbalance_value", None)
            row_data = [
                getattr(blade, "serial_number", "") if blade else "",
                getattr(slot, "slot_number", ""),
                getattr(slot, "position", ""),
                "Yes" if getattr(slot, "is_balanced", False) else "No",
                round(float(unbal), 4) if unbal is not None else "",
                getattr(slot, "balancing_remarks", "") or "",
                _fmt_dt(getattr(slot, "allocated_at", None)),
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws_slots.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border()
                if alt_fill(row_idx):
                    cell.fill = alt_fill(row_idx)
        auto_width(ws_slots)

        # ---- Sheet 4: Workflow History ----
        ws_wf = wb.create_sheet("Workflow History")
        wf_headers = [
            "Serial Number", "From Status", "To Status",
            "Actor ID", "Timestamp", "Remarks",
        ]
        write_headers(ws_wf, wf_headers)
        for row_idx, event in enumerate(workflow_events, start=2):
            blade = blade_by_id_xl.get(str(getattr(event, "blade_id", "")))
            row_data = [
                getattr(blade, "serial_number", "") if blade else "",
                _ev(getattr(event, "from_status", None)),
                _ev(getattr(event, "to_status",   None)),
                str(getattr(event, "action_by_id", "") or "")[:36],
                _fmt_dt(getattr(event, "timestamp", None)),
                getattr(event, "remarks", "") or "",
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws_wf.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border()
                if alt_fill(row_idx):
                    cell.fill = alt_fill(row_idx)
        auto_width(ws_wf)

        # ---- Sheet 5: Work Order Traceability ----
        ws_bt = wb.create_sheet("Work Order Traceability")
        bt_headers = [
            "Work Order Number", "Serial Number", "Melt Number",
            "Blade Type", "Status", "Slot Number",
            "Rocking Value", "Creep Value",
        ]
        write_headers(ws_bt, bt_headers)

        # Build lookup maps: blade_id → slot_number and blade_id → (rocking, creep)
        slot_by_blade_xl = {str(getattr(s, "blade_id", "")): getattr(s, "slot_number", "") for s in slots}
        meas_by_blade_xl: dict[str, Any] = {}
        for m in measurements:
            bid = str(getattr(m, "blade_id", ""))
            rv = getattr(m, "rocking_value", None)
            cv = getattr(m, "creep_value", None)
            existing = meas_by_blade_xl.get(bid)
            # Keep the most recent entry that has rocking/creep data
            if existing is None or (rv is not None or cv is not None):
                meas_by_blade_xl[bid] = (rv, cv)

        # Sort blades by work_order_number then serial_number for clean grouping
        sorted_blades = sorted(
            blades,
            key=lambda b: (getattr(b, "work_order_number", "") or "", getattr(b, "serial_number", "")),
        )
        for row_idx, blade in enumerate(sorted_blades, start=2):
            bid = str(getattr(blade, "id", ""))
            rv, cv = meas_by_blade_xl.get(bid, (None, None))
            row_data = [
                getattr(blade, "work_order_number", "") or "—",
                getattr(blade, "serial_number", ""),
                getattr(blade, "melt_number", ""),
                _ev(getattr(blade, "blade_type", "")),
                _ev(getattr(blade, "status", "")),
                slot_by_blade_xl.get(bid, "—"),
                round(float(rv), 4) if rv is not None else "",
                round(float(cv), 4) if cv is not None else "",
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws_bt.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border()
                if alt_fill(row_idx):
                    cell.fill = alt_fill(row_idx)
        auto_width(ws_bt)

        buf = io.BytesIO()
        wb.save(buf)
        logger.info(
            "excel_report_generated",
            blade_count=len(blades),
            sheet_count=len(wb.sheetnames),
        )
        return buf.getvalue()

    # ==================================================================
    # PDF report
    # ==================================================================

    async def generate_blade_report_pdf(
        self,
        blade_ids: list[UUID],
        db: AsyncSession,
        filter_params: dict | None = None,
    ) -> bytes:
        """
        Build a multi-page PDF blade report using ReportLab.

        Structure
        ---------
        * Company header with report title and generated timestamp
        * Blade detail table (one section per blade)
        * Measurements table
        * Workflow timeline
        * Page numbers in footer

        Returns
        -------
        bytes
            Raw PDF bytes.
        """
        try:
            from reportlab.lib import colors  # type: ignore[import]
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                HRFlowable,
                PageBreak,
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError as exc:
            raise RuntimeError("reportlab is required for PDF reports.") from exc

        blades, measurements, slots, workflow_events = await self._fetch_report_data(
            blade_ids, db, filter_params
        )

        buf = io.BytesIO()
        generated_at = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        # ── Page setup ────────────────────────────────────────────────────
        # A4 = 21 cm wide. With 1.5 cm left+right margins, usable = 18 cm
        LEFT_M = RIGHT_M = 1.5 * cm
        TOP_M = 2.0 * cm
        BOT_M = 2.0 * cm
        PAGE_W = A4[0] - LEFT_M - RIGHT_M  # ~18 cm

        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=LEFT_M,
            rightMargin=RIGHT_M,
            topMargin=TOP_M,
            bottomMargin=BOT_M,
            title="Blade Rocking & Creep Test Report",
            author="Blade Management System",
        )

        # ── Custom styles ─────────────────────────────────────────────────
        from reportlab.lib.styles import ParagraphStyle  # type: ignore[import]
        from reportlab.lib.enums import TA_LEFT, TA_CENTER  # type: ignore[import]

        base = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "ReportTitle2",
            parent=base["Normal"],
            fontSize=16,
            leading=20,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1F4E79"),
            spaceAfter=4,
        )
        subtitle_style = ParagraphStyle(
            "ReportSubtitle2",
            parent=base["Normal"],
            fontSize=9,
            leading=13,
            textColor=colors.HexColor("#444444"),
            spaceAfter=8,
        )
        section_style = ParagraphStyle(
            "SectionHeading2",
            parent=base["Normal"],
            fontSize=11,
            leading=14,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1F4E79"),
            spaceBefore=10,
            spaceAfter=4,
        )
        # Cell styles — DO NOT set wordWrap="CJK" (breaks English text splitting).
        # splitLongWords forces mid-word breaks for tokens with no spaces (dates, serials).
        # leading=12 for 8pt font gives 1.5× line spacing — critical to prevent overlap.
        cell_style = ParagraphStyle(
            "CellText2",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=12,
            splitLongWords=1,
            allowWidows=0,
        )
        cell_hdr = ParagraphStyle(
            "CellHeader2",
            parent=base["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=12,
            textColor=colors.white,
            alignment=TA_CENTER,
            splitLongWords=1,
        )

        def P(text: str, center: bool = False) -> Paragraph:
            """Wrap any value as a Paragraph — ensures rows auto-size to content."""
            s = str(text) if text not in (None, "", "None") else "—"
            st = ParagraphStyle(
                "C2", parent=cell_style,
                alignment=TA_CENTER if center else TA_LEFT,
            )
            return Paragraph(s, st)

        def H(text: str) -> Paragraph:
            return Paragraph(str(text), cell_hdr)

        DARK_BLUE = colors.HexColor("#1F4E79")
        MED_BLUE  = colors.HexColor("#2E75B6")
        ALT_ROW   = colors.HexColor("#EEF4FA")
        BORDER    = colors.HexColor("#BBBBBB")

        def base_ts(header_color: Any = DARK_BLUE) -> list:
            # NOTE: no FONTSIZE/LEADING here — those only apply to string cells,
            # not Paragraph cells, and can cause row-height miscalculation.
            return [
                ("BACKGROUND",    (0, 0), (-1, 0),  header_color),
                ("LINEBELOW",     (0, 0), (-1, 0),  1.0, header_color),
                ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, ALT_ROW]),
                ("GRID",          (0, 0), (-1, -1), 0.4, BORDER),
                ("VALIGN",        (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING",   (0, 0), (-1, -1), 6),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ]

        story: list[Any] = []

        # ── Header ────────────────────────────────────────────────────────
        story.append(Paragraph("Blade Rocking &amp; Creep Test Management System", title_style))
        story.append(Paragraph(
            f"Blade Inspection Report &nbsp;&nbsp;&bull;&nbsp;&nbsp; "
            f"Generated: <b>{generated_at}</b> &nbsp;&nbsp;&bull;&nbsp;&nbsp; "
            f"Total blades: <b>{len(blades)}</b>",
            subtitle_style,
        ))
        story.append(HRFlowable(width="100%", thickness=1.5, color=DARK_BLUE, spaceAfter=8))

        # blade_id → blade lookup shared across all sections
        blade_by_id = {str(getattr(b, "id", "")): b for b in blades}

        # ── 1. Blade Summary ─────────────────────────────────────────────
        # 5 columns — generous widths so status labels ("Balancing In Progress") fit
        # Total = 4.5+3.5+3.5+3.5+3.0 = 18.0 cm
        story.append(Paragraph("1. Blade Summary", section_style))
        C_SUM = [4.5*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.0*cm]
        sum_rows = [[H("Serial Number"), H("Work Order"), H("Part Number"), H("Status"), H("Type / Created")]]
        for b in blades:
            btype   = _ev(getattr(b, "blade_type", "LPTR")) or "LPTR"
            created = _fmt_dt(getattr(b, "created_at", None))
            status  = _ev(getattr(b, "status", "")).replace("_", " ").title()
            sum_rows.append([
                P(getattr(b, "serial_number", "")),
                P(getattr(b, "work_order_number", "")),
                P(getattr(b, "part_number", "")),
                P(status),
                P(f"{btype}  {created}"),
            ])
        tbl = Table(sum_rows, colWidths=C_SUM, repeatRows=1)
        tbl.setStyle(TableStyle(base_ts(DARK_BLUE)))
        story.append(tbl)
        story.append(Spacer(1, 0.5 * cm))

        # ── 2. Measurements ──────────────────────────────────────────────
        # 7 columns — no Melt col (saves space); date column 3.2 cm to safely
        # hold "YYYY-MM-DD HH:MM" (16 chars) after word-wrap at the space.
        # Total = 3.5+2.0+2.5+2.5+2.3+2.0+3.2 = 18.0 cm
        if measurements:
            story.append(Paragraph("2. Measurements", section_style))
            C_MEAS = [3.5*cm, 2.0*cm, 2.5*cm, 2.5*cm, 2.3*cm, 2.0*cm, 3.2*cm]
            meas_rows = [[
                H("Serial No."), H("Type"),
                H("Weight (g)"), H("Static Moment (g.cm)"),
                H("Rocking"), H("Creep"), H("Measured At"),
            ]]
            for m in measurements:
                blade = blade_by_id.get(str(getattr(m, "blade_id", "")))
                serial = getattr(blade, "serial_number", "") if blade else ""
                w  = getattr(m, "weight_grams", None)
                sm = getattr(m, "static_moment_gcm", None)
                rv = getattr(m, "rocking_value", None)
                cv = getattr(m, "creep_value", None)
                meas_rows.append([
                    P(serial),
                    P(_ev(getattr(m, "measurement_type", "")).title()),
                    P(f"{float(w):.2f}"  if w  is not None else "—", center=True),
                    P(f"{float(sm):.2f}" if sm is not None else "—", center=True),
                    P(f"{float(rv):.4f}" if rv is not None else "—", center=True),
                    P(f"{float(cv):.4f}" if cv is not None else "—", center=True),
                    P(_fmt_dt(getattr(m, "measured_at", None))),
                ])
            tbl = Table(meas_rows, colWidths=C_MEAS, repeatRows=1)
            tbl.setStyle(TableStyle(base_ts(MED_BLUE)))
            story.append(tbl)
            story.append(Spacer(1, 0.5 * cm))

        # ── 3. Slot Allocations ──────────────────────────────────────────
        # 5 columns — Remarks gets the extra space.
        # Total = 4.0+2.5+2.5+2.5+6.5 = 18.0 cm
        if slots:
            story.append(Paragraph("3. Slot Allocations", section_style))
            C_SLOT = [4.0*cm, 2.5*cm, 2.5*cm, 2.5*cm, 6.5*cm]
            slot_rows = [[H("Serial No."), H("Slot No."), H("Balanced"), H("Imbalance"), H("Remarks")]]
            for s in slots:
                blade = blade_by_id.get(str(getattr(s, "blade_id", "")))
                serial = getattr(blade, "serial_number", "") if blade else ""
                unbal  = getattr(s, "unbalance_value", None)
                slot_rows.append([
                    P(serial),
                    P(str(getattr(s, "slot_number", "")), center=True),
                    P("Yes" if getattr(s, "is_balanced", False) else "No", center=True),
                    P(f"{float(unbal):.4f}" if unbal is not None else "—", center=True),
                    P(str(getattr(s, "balancing_remarks", "") or "")),
                ])
            tbl = Table(slot_rows, colWidths=C_SLOT, repeatRows=1)
            tbl.setStyle(TableStyle(base_ts(colors.HexColor("#2C7A52"))))
            story.append(tbl)
            story.append(Spacer(1, 0.5 * cm))

        # ── 4. Workflow Timeline ─────────────────────────────────────────
        # 4 columns — From/To status each get 4 cm, plenty for "Balancing In Progress"
        # Total = 4.0+4.0+4.0+6.0 = 18.0 cm
        if workflow_events:
            story.append(Paragraph("4. Workflow Timeline", section_style))
            C_WF = [4.0*cm, 4.0*cm, 4.0*cm, 6.0*cm]
            wf_rows = [[H("Serial No."), H("From Status"), H("To Status"), H("Timestamp")]]
            for ev in workflow_events:
                blade = blade_by_id.get(str(getattr(ev, "blade_id", "")))
                serial    = getattr(blade, "serial_number", "") if blade else ""
                from_stat = (_ev(getattr(ev, "from_status", None)) or "—").replace("_", " ").title()
                to_stat   = (_ev(getattr(ev, "to_status",   None)) or "—").replace("_", " ").title()
                wf_rows.append([
                    P(serial),
                    P(from_stat),
                    P(to_stat),
                    P(_fmt_dt(getattr(ev, "timestamp", None))),
                ])
            tbl = Table(wf_rows, colWidths=C_WF, repeatRows=1)
            tbl.setStyle(TableStyle(base_ts(DARK_BLUE)))
            story.append(tbl)

        # ── 5. Work Order Traceability ────────────────────────────────────
        # 8 cols: Work Order(2.5) Serial(3.5) Melt(2.5) Type(1.5) Status(3.0) Slot(1.5) Rock(2.0) Creep(1.5) = 18.0
        story.append(Paragraph("5. Work Order Traceability", section_style))
        C_BT = [2.5*cm, 3.5*cm, 2.5*cm, 1.5*cm, 3.0*cm, 1.5*cm, 2.0*cm, 1.5*cm]
        bt_rows = [[
            H("Work Order"), H("Serial No."), H("Melt No."),
            H("Type"), H("Status"), H("Slot"),
            H("Rocking"), H("Creep"),
        ]]

        slot_by_blade_pdf = {str(getattr(s, "blade_id", "")): getattr(s, "slot_number", "") for s in slots}
        meas_by_blade_pdf: dict[str, Any] = {}
        for m in measurements:
            bid = str(getattr(m, "blade_id", ""))
            rv = getattr(m, "rocking_value", None)
            cv = getattr(m, "creep_value", None)
            existing = meas_by_blade_pdf.get(bid)
            if existing is None or (rv is not None or cv is not None):
                meas_by_blade_pdf[bid] = (rv, cv)

        sorted_blades_pdf = sorted(
            blades,
            key=lambda b: (getattr(b, "work_order_number", "") or "", getattr(b, "serial_number", "")),
        )
        for b in sorted_blades_pdf:
            bid = str(getattr(b, "id", ""))
            rv, cv = meas_by_blade_pdf.get(bid, (None, None))
            status_txt = (_ev(getattr(b, "status", "")) or "—").replace("_", " ").title()
            bt_rows.append([
                P(getattr(b, "work_order_number", "") or "—"),
                P(getattr(b, "serial_number", "")),
                P(getattr(b, "melt_number", "")),
                P(_ev(getattr(b, "blade_type", "")) or "—", center=True),
                P(status_txt),
                P(str(slot_by_blade_pdf.get(bid, "—")), center=True),
                P(f"{float(rv):.4f}" if rv is not None else "—", center=True),
                P(f"{float(cv):.4f}" if cv is not None else "—", center=True),
            ])
        tbl = Table(bt_rows, colWidths=C_BT, repeatRows=1)
        tbl.setStyle(TableStyle(base_ts(colors.HexColor("#6B3FA0"))))
        story.append(tbl)
        story.append(Spacer(1, 0.5 * cm))

        # ── Page footer (number + date) ───────────────────────────────────
        def _page_footer(canvas: Any, doc: Any) -> None:
            canvas.saveState()
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.HexColor("#888888"))
            y = 1.2 * cm
            # Left: company name
            canvas.drawString(LEFT_M, y, "Blade Rocking & Creep Test Management System")
            # Centre: report title
            canvas.drawCentredString(A4[0] / 2, y, "Confidential — Internal Use Only")
            # Right: page number
            canvas.drawRightString(
                A4[0] - RIGHT_M, y,
                f"Page {canvas.getPageNumber()}  |  {generated_at}",
            )
            # Top border above footer
            canvas.setStrokeColor(colors.HexColor("#CCCCCC"))
            canvas.setLineWidth(0.5)
            canvas.line(LEFT_M, y + 0.4 * cm, A4[0] - RIGHT_M, y + 0.4 * cm)
            canvas.restoreState()

        doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
        logger.info("pdf_report_generated", blade_count=len(blades))
        return buf.getvalue()

    # ==================================================================
    # Batch (work order) report — slot/serial/melt/weight/static moment/
    # rocking/creep, one sheet or table, columns adapted to blade_type.
    # ==================================================================

    async def generate_batch_report_excel(
        self,
        work_order_number: str,
        db: AsyncSession,
    ) -> bytes:
        """
        Build a single-sheet Excel batch report for *work_order_number*.

        Columns: Slot No., Serial No., Melt No., Weight (g),
        Static Moment (g.cm), Rocking — plus Creep for LPTR work orders
        only (HPTR blades never undergo creep testing).

        Raises
        ------
        ValueError
            If the work order doesn't exist or has no blades.
        RuntimeError
            If openpyxl isn't installed.
        """
        try:
            import openpyxl  # type: ignore[import]
            from openpyxl.styles import Alignment, Border, Font, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for Excel reports.") from exc

        work_order, rows = await self._fetch_batch_report_rows(work_order_number, db)
        blade_type = _ev(work_order.blade_type)
        is_lptr = blade_type == "LPTR"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{blade_type} Batch Report"

        headers = ["Slot No.", "Serial No.", "Melt No.", "Weight (g)", "Static Moment (g·cm)", "Rocking"]
        if is_lptr:
            headers.append("Creep")
        n_cols = len(headers)

        # ── Styling matched to the original HPTR_*.xls shop-floor sheets,
        # and structured the same way as the PDF version: a two-column-block
        # layout (first half of the rows on the left, remainder on the right
        # — a plain page-fitting split, not a physical W1/W2 split) so the
        # whole sheet — not just the print scaling — reads as one page's
        # worth of content instead of one long 90-row column. ──────────────
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=10)
        label_font = Font(bold=True, size=9)
        data_font = Font(size=9)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        left_col_start, right_col_start = 1, n_cols + 2  # one blank spacer column between blocks
        total_cols = right_col_start + n_cols - 1

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=f"Work Order {work_order_number} — {blade_type} Batch Report")
        title_cell.font = title_font
        title_cell.alignment = center
        ws.row_dimensions[1].height = 20

        generated_at = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        subtitle_cell = ws.cell(row=2, column=1, value=f"Generated: {generated_at}  •  {len(rows)} blade(s)")
        subtitle_cell.font = Font(italic=True, size=8, color="666666")
        subtitle_cell.alignment = center

        # Info fields condensed onto a single row (rather than 7 stacked
        # rows) — the same vertical-space saving used in the PDF, so the
        # two 45-row blocks below have room to sit on one printed page.
        info_fields = [
            ("Shop Order No.", work_order.shop_order_number),
            ("Part No.", work_order.part_number),
            ("Engine No.", work_order.engine_number or "—"),
            ("Engine Hours", work_order.engine_hours),
            ("Component Hours", work_order.component_hours or "—"),
            ("Blade Type", blade_type),
        ]
        info_row = 3
        ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=total_cols)
        info_text = "   |   ".join(f"{label}: {value}" for label, value in info_fields)
        info_cell = ws.cell(row=info_row, column=1, value=info_text)
        info_cell.font = label_font
        info_cell.alignment = center

        header_row = info_row + 2

        def _write_block(col_start: int, block_rows: list[dict]) -> None:
            for i, h in enumerate(headers):
                cell = ws.cell(row=header_row, column=col_start + i, value=h)
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            for row_offset, row in enumerate(block_rows, start=1):
                slot = row["slot_number"]
                w = row["weight_grams"]
                sm = row["static_moment_gcm"]
                rv = row["rocking_value"]
                cv = row["creep_value"]
                values = [
                    int(slot) if slot is not None and str(slot).isdigit() else (slot or "—"),
                    row["serial_number"],
                    row["melt_number"],
                    float(w) if w is not None else None,
                    float(sm) if sm is not None else None,
                    float(rv) if rv is not None else None,
                ]
                if is_lptr:
                    values.append(float(cv) if cv is not None else None)
                r = header_row + row_offset
                for i, value in enumerate(values):
                    cell = ws.cell(row=r, column=col_start + i, value=value)
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = left if i in (1, 2) else center
                    if i in (3, 4):  # Weight (g), Static Moment (g·cm)
                        cell.number_format = "0.00"
                    elif i in (5, 6):  # Rocking, Creep
                        cell.number_format = "0.0000"

        midpoint = (len(rows) + 1) // 2
        _write_block(left_col_start, rows[:midpoint])
        _write_block(right_col_start, rows[midpoint:])

        col_widths = [7, 9, 11, 9, 12, 9]
        if is_lptr:
            col_widths.append(9)
        for i, width in enumerate(col_widths):
            ws.column_dimensions[ws.cell(row=header_row, column=left_col_start + i).column_letter].width = width
            ws.column_dimensions[ws.cell(row=header_row, column=right_col_start + i).column_letter].width = width
        ws.column_dimensions[ws.cell(row=header_row, column=left_col_start + n_cols).column_letter].width = 3
        ws.freeze_panes = f"A{header_row + 1}"

        # Print setup: scale the whole sheet to one A4 page when printed —
        # doesn't affect the on-screen/editing view, only the print output.
        last_row = header_row + midpoint
        last_col_letter = get_column_letter(total_cols)
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = ws.page_margins.right = 0.3
        ws.page_margins.top = ws.page_margins.bottom = 0.3
        ws.page_margins.header = ws.page_margins.footer = 0.1
        ws.print_area = f"A1:{last_col_letter}{last_row}"

        buf = io.BytesIO()
        wb.save(buf)
        logger.info(
            "batch_report_excel_generated",
            work_order_number=work_order_number,
            blade_type=blade_type,
            row_count=len(rows),
        )
        return buf.getvalue()

    async def generate_batch_report_pdf(
        self,
        work_order_number: str,
        db: AsyncSession,
    ) -> bytes:
        """
        Build a single-table PDF batch report for *work_order_number*, with
        the same columns as :meth:`generate_batch_report_excel`.
        """
        try:
            from reportlab.lib import colors  # type: ignore[import]
            from reportlab.lib.enums import TA_CENTER, TA_LEFT  # type: ignore[import]
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet  # type: ignore[import]
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError as exc:
            raise RuntimeError("reportlab is required for PDF reports.") from exc

        work_order, rows = await self._fetch_batch_report_rows(work_order_number, db)
        blade_type = _ev(work_order.blade_type)
        is_lptr = blade_type == "LPTR"
        generated_at = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        # Landscape + a two-column-block layout (first half of the rows on
        # the left, remainder on the right — a plain split for page-fitting,
        # not a physical W1/W2 balancing split) is what makes a fixed 90-row
        # batch fit on a single printed A4 page at a still-legible font size;
        # a single 90-row column would need ~30 pages at a readable row height.
        PAGE_SIZE = landscape(A4)
        buf = io.BytesIO()
        LEFT_M = RIGHT_M = 0.8 * cm
        doc = SimpleDocTemplate(
            buf,
            pagesize=PAGE_SIZE,
            leftMargin=LEFT_M,
            rightMargin=RIGHT_M,
            topMargin=0.6 * cm,
            bottomMargin=1.3 * cm,
            title=f"Work Order {work_order_number} Batch Report",
            author="Blade Management System",
        )

        # ── Styling matched to the original HPTR_*.xls shop-floor sheets:
        # bold centered title (plain black, no colour); plain bold headers
        # with no fill, thin borders; bordered label/value info table with
        # no alternating row shading. Font sizes and paddings are kept small
        # throughout specifically so the two 45-row blocks fit in the space
        # left after the title/info block on one page. ────────────────────
        base = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "BatchTitle", parent=base["Normal"], fontSize=12, leading=14,
            fontName="Helvetica-Bold", textColor=colors.black, alignment=TA_CENTER, spaceAfter=1,
        )
        info_style = ParagraphStyle(
            "BatchInfo", parent=base["Normal"], fontSize=7.5, leading=10,
            textColor=colors.black, alignment=TA_CENTER,
        )
        cell_style = ParagraphStyle(
            "BatchCell", parent=base["Normal"], fontName="Helvetica", fontSize=6.2,
            leading=7.5, splitLongWords=1, allowWidows=0,
        )
        cell_hdr = ParagraphStyle(
            "BatchCellHdr", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=6.8,
            leading=8, textColor=colors.black, alignment=TA_CENTER, splitLongWords=1,
        )

        def P(text: Any, center: bool = False) -> Paragraph:
            s = str(text) if text not in (None, "", "None") else "—"
            st = ParagraphStyle("BC", parent=cell_style, alignment=TA_CENTER if center else TA_LEFT)
            return Paragraph(s, st)

        def H(text: str) -> Paragraph:
            return Paragraph(str(text), cell_hdr)

        BORDER = colors.black

        story: list[Any] = []
        story.append(Paragraph(f"Work Order {work_order_number} — {blade_type} Batch Report", title_style))

        # Generated/total-blades and the info fields condensed onto a single
        # line each (rather than a subtitle line plus 7 stacked info rows) —
        # the biggest vertical-space saving that makes room for the
        # two-block table below to fit on one page.
        story.append(Paragraph(
            f"Generated: <b>{generated_at}</b> &nbsp;&nbsp;&bull;&nbsp;&nbsp; Total blades: <b>{len(rows)}</b>",
            info_style,
        ))
        info_fields = [
            ("Shop Order No.", work_order.shop_order_number),
            ("Part No.", work_order.part_number),
            ("Engine No.", work_order.engine_number or "—"),
            ("Engine Hours", work_order.engine_hours),
            ("Component Hours", work_order.component_hours or "—"),
            ("Blade Type", blade_type),
        ]
        info_line = "&nbsp;&nbsp;|&nbsp;&nbsp;".join(
            f"<b>{label}:</b> {value}" for label, value in info_fields
        )
        story.append(Paragraph(info_line, info_style))
        story.append(Spacer(1, 0.15 * cm))

        headers = [H("Slot"), H("Serial"), H("Melt No."), H("Wt (g)"), H("Static Mt (g·cm)"), H("Rocking")]
        col_widths = [1.2 * cm, 2.0 * cm, 2.3 * cm, 2.0 * cm, 2.5 * cm, 2.0 * cm]
        if is_lptr:
            headers.append(H("Creep"))
            col_widths = [1.1 * cm, 1.8 * cm, 2.0 * cm, 1.8 * cm, 2.2 * cm, 1.8 * cm, 1.8 * cm]

        def _block_table(block_rows: list[dict]) -> Table:
            block_data: list[list[Any]] = [headers]
            for row in block_rows:
                slot = row["slot_number"]
                w = row["weight_grams"]
                sm = row["static_moment_gcm"]
                rv = row["rocking_value"]
                cv = row["creep_value"]
                line = [
                    P(slot if slot is not None else "—", center=True),
                    P(row["serial_number"], center=True),
                    P(row["melt_number"], center=True),
                    P(f"{float(w):.2f}" if w is not None else "—", center=True),
                    P(f"{float(sm):.2f}" if sm is not None else "—", center=True),
                    P(f"{float(rv):.4f}" if rv is not None else "—", center=True),
                ]
                if is_lptr:
                    line.append(P(f"{float(cv):.4f}" if cv is not None else "—", center=True))
                block_data.append(line)

            block_tbl = Table(block_data, colWidths=col_widths, repeatRows=1)
            block_tbl.setStyle(TableStyle([
                ("LINEBELOW", (0, 0), (-1, 0), 1.0, BORDER),
                ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 1.0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.0),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ]))
            return block_tbl

        midpoint = (len(rows) + 1) // 2
        left_rows, right_rows = rows[:midpoint], rows[midpoint:]

        outer = Table(
            [[_block_table(left_rows), "", _block_table(right_rows)]],
            colWidths=[sum(col_widths), 0.6 * cm, sum(col_widths)],
        )
        outer.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(outer)

        def _page_footer(canvas: Any, doc: Any) -> None:
            canvas.saveState()
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.HexColor("#888888"))
            y = 1.0 * cm
            canvas.drawString(LEFT_M, y, "Blade Rocking & Creep Test Management System")
            canvas.drawCentredString(PAGE_SIZE[0] / 2, y, "Confidential — Internal Use Only")
            canvas.drawRightString(PAGE_SIZE[0] - RIGHT_M, y, f"Page {canvas.getPageNumber()}  |  {generated_at}")
            canvas.setStrokeColor(colors.HexColor("#CCCCCC"))
            canvas.setLineWidth(0.5)
            canvas.line(LEFT_M, y + 0.35 * cm, PAGE_SIZE[0] - RIGHT_M, y + 0.35 * cm)
            canvas.restoreState()

        doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
        logger.info(
            "batch_report_pdf_generated",
            work_order_number=work_order_number,
            blade_type=blade_type,
            row_count=len(rows),
        )
        return buf.getvalue()

    async def _fetch_batch_report_rows(
        self,
        work_order_number: str,
        db: AsyncSession,
    ) -> tuple[Any, list[dict]]:
        """
        Fetch one row per blade in *work_order_number*: active slot number,
        serial number, melt number, and weight/static moment/rocking/creep
        from each blade's most recent measurement.

        Rows are sorted by numeric slot number where assigned, falling back
        to serial number for unassigned blades.

        Returns
        -------
        tuple[WorkOrder, list[dict]]

        Raises
        ------
        ValueError
            If the work order doesn't exist or has no blades.
        """
        from app.models.blade import Blade
        from app.models.measurement import Measurement
        from app.models.slot_allocation import SlotAllocation
        from app.models.work_order import WorkOrder

        work_order = (
            await db.execute(select(WorkOrder).where(WorkOrder.work_order_number == work_order_number))
        ).scalar_one_or_none()
        if work_order is None:
            raise ValueError(f"Work Order '{work_order_number}' not found")

        blades = list(
            (
                await db.execute(
                    select(Blade)
                    .where(Blade.work_order_number == work_order_number, Blade.deleted_at.is_(None))
                    .order_by(Blade.serial_number)
                )
            )
            .scalars()
            .all()
        )
        if not blades:
            raise ValueError(f"Work Order '{work_order_number}' has no blades")

        blade_ids = [b.id for b in blades]

        slot_rows = (
            await db.execute(
                select(SlotAllocation.blade_id, SlotAllocation.slot_number).where(
                    SlotAllocation.blade_id.in_(blade_ids), SlotAllocation.is_active.is_(True)
                )
            )
        ).all()
        slot_map = {r.blade_id: r.slot_number for r in slot_rows}

        subq = (
            select(Measurement.blade_id, func.max(Measurement.measured_at).label("latest_at"))
            .where(Measurement.blade_id.in_(blade_ids))
            .group_by(Measurement.blade_id)
            .subquery()
        )
        meas_rows = (
            await db.execute(
                select(
                    Measurement.blade_id,
                    Measurement.weight_grams,
                    Measurement.static_moment_gcm,
                    Measurement.rocking_value,
                    Measurement.creep_value,
                ).join(
                    subq,
                    (Measurement.blade_id == subq.c.blade_id) & (Measurement.measured_at == subq.c.latest_at),
                )
            )
        ).all()
        meas_map = {r.blade_id: r for r in meas_rows}

        def _slot_sort_key(blade: Any) -> tuple:
            slot = slot_map.get(blade.id)
            if slot is not None and str(slot).isdigit():
                return (0, int(slot))
            return (1, blade.serial_number)

        rows: list[dict] = []
        for blade in sorted(blades, key=_slot_sort_key):
            meas = meas_map.get(blade.id)
            rows.append({
                "slot_number": slot_map.get(blade.id),
                "serial_number": blade.serial_number,
                "melt_number": blade.melt_number,
                "weight_grams": meas.weight_grams if meas else None,
                "static_moment_gcm": meas.static_moment_gcm if meas else None,
                "rocking_value": meas.rocking_value if meas else None,
                "creep_value": meas.creep_value if meas else None,
            })

        return work_order, rows

    # ==================================================================
    # Dashboard summary
    # ==================================================================

    async def generate_dashboard_summary(self, db: AsyncSession) -> dict:
        """
        Compute aggregated statistics for dashboard summary cards.

        Returns
        -------
        dict
            Keys:

            * ``total_blades`` — total blade count
            * ``blades_by_status`` — ``{status: count}``
            * ``blades_by_station`` — ``{station: count}``
            * ``rejection_rate_pct`` — percentage of blades rejected
            * ``avg_processing_hours`` — mean hours from receipt to completion
            * ``generated_at`` — ISO-8601 UTC timestamp string
        """
        from app.models.blade import Blade  # late import to avoid circulars

        try:
            # Total count
            total_result = await db.execute(select(func.count()).select_from(Blade))
            total: int = total_result.scalar_one() or 0

            # Count by status
            status_result = await db.execute(
                select(Blade.status, func.count(Blade.id)).group_by(Blade.status)
            )
            blades_by_status: dict[str, int] = {
                str(row[0]): row[1] for row in status_result.all()
            }

            # Count by station (use current_station_id column, not the relationship)
            station_result = await db.execute(
                select(Blade.current_station_id, func.count(Blade.id)).group_by(
                    Blade.current_station_id
                )
            )
            blades_by_station: dict[str, int] = {
                str(row[0] or "Unassigned"): row[1] for row in station_result.all()
            }

            # Rejection rate
            rejected: int = blades_by_status.get("REJECTED", 0)
            rejection_rate = round((rejected / total * 100), 2) if total else 0.0

            # Average processing time (hours) for COMPLETED blades
            avg_hours: float = 0.0
            try:
                from sqlalchemy import cast, Float, extract

                avg_result = await db.execute(
                    select(
                        func.avg(
                            extract(
                                "epoch",
                                Blade.updated_at - Blade.created_at,
                            )
                        )
                    ).where(Blade.status == "COMPLETED")
                )
                avg_seconds = avg_result.scalar_one()
                if avg_seconds is not None:
                    avg_hours = round(float(avg_seconds) / 3600, 2)
            except Exception:  # noqa: BLE001
                avg_hours = 0.0

        except Exception as exc:  # noqa: BLE001
            logger.error("dashboard_summary_query_failed", error=str(exc))
            return {
                "total_blades": 0,
                "blades_by_status": {},
                "blades_by_station": {},
                "rejection_rate_pct": 0.0,
                "avg_processing_hours": 0.0,
                "generated_at": datetime.now(tz=timezone.utc).isoformat(),
                "error": str(exc),
            }

        return {
            "total_blades": total,
            "blades_by_status": blades_by_status,
            "blades_by_station": blades_by_station,
            "rejection_rate_pct": rejection_rate,
            "avg_processing_hours": avg_hours,
            "generated_at": datetime.now(tz=timezone.utc).isoformat(),
        }

    # ==================================================================
    # Private data-fetching helpers
    # ==================================================================

    async def _fetch_report_data(
        self,
        blade_ids: list[UUID],
        db: AsyncSession,
        filter_params: dict | None = None,
    ) -> tuple[list, list, list, list]:
        """
        Fetch all data required for a blade report from the database.

        Returns a 4-tuple: (blades, measurements, slots, workflow_events).
        Returns empty lists gracefully if the models/tables don't exist yet.
        """
        blades: list = []
        measurements: list = []
        slots: list = []
        workflow_events: list = []

        try:
            from datetime import timezone
            from sqlalchemy import and_
            from app.models.blade import Blade
            from app.models.enums import BladeStatus as BladeStatusEnum

            if blade_ids:
                res = await db.execute(select(Blade).where(Blade.id.in_(blade_ids)))
            else:
                fp = filter_params or {}
                conditions: list = [Blade.deleted_at.is_(None)]

                status_list = fp.get("status")
                if status_list:
                    conditions.append(
                        Blade.status.in_([BladeStatusEnum(s) for s in status_list])
                    )
                elif not fp.get("include_rejected", True):
                    conditions.append(Blade.status != BladeStatusEnum.REJECTED)

                date_from = fp.get("date_from")
                if date_from:
                    from datetime import datetime
                    dt_from = datetime.fromisoformat(str(date_from)).replace(tzinfo=timezone.utc)
                    conditions.append(Blade.created_at >= dt_from)

                date_to = fp.get("date_to")
                if date_to:
                    from datetime import datetime
                    dt_to = datetime.fromisoformat(str(date_to)).replace(
                        hour=23, minute=59, second=59, microsecond=999999, tzinfo=timezone.utc
                    )
                    conditions.append(Blade.created_at <= dt_to)

                serial_number = fp.get("serial_number")
                if serial_number:
                    conditions.append(Blade.serial_number.ilike(f"%{serial_number}%"))

                part_number = fp.get("part_number")
                if part_number:
                    conditions.append(Blade.part_number.ilike(f"%{part_number}%"))

                station_ids = fp.get("station_ids")
                if station_ids:
                    from uuid import UUID as _UUID
                    conditions.append(
                        Blade.current_station_id.in_([_UUID(str(s)) for s in station_ids])
                    )

                res = await db.execute(
                    select(Blade).where(and_(*conditions)).order_by(Blade.created_at.desc())
                )

            blades = list(res.scalars().all())
            blade_ids = [b.id for b in blades]
        except Exception as exc:  # noqa: BLE001
            logger.warning("report_blade_fetch_failed", error=str(exc))

        try:
            from app.models.measurement import Measurement

            if blade_ids:
                res = await db.execute(
                    select(Measurement).where(Measurement.blade_id.in_(blade_ids))
                )
                measurements = list(res.scalars().all())
        except Exception as exc:  # noqa: BLE001
            logger.warning("report_measurement_fetch_failed", error=str(exc))

        try:
            from app.models.slot_allocation import SlotAllocation

            if blade_ids:
                res = await db.execute(
                    select(SlotAllocation)
                    .where(SlotAllocation.blade_id.in_(blade_ids),
                           SlotAllocation.is_active.is_(True))
                )
                slots = list(res.scalars().all())
        except Exception as exc:  # noqa: BLE001
            logger.warning("report_slot_fetch_failed", error=str(exc))

        try:
            from app.models.workflow import WorkflowLog

            if blade_ids:
                res = await db.execute(
                    select(WorkflowLog)
                    .where(WorkflowLog.blade_id.in_(blade_ids))
                    .order_by(WorkflowLog.timestamp.asc())
                )
                workflow_events = list(res.scalars().all())
        except Exception as exc:  # noqa: BLE001
            logger.warning("report_workflow_fetch_failed", error=str(exc))

        return blades, measurements, slots, workflow_events


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _fmt_dt(value: Any) -> str:
    """Format a datetime to a human-readable string, or return empty string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def _ev(value: Any) -> str:
    """Return the plain string value of a Python enum or any other value."""
    if value is None:
        return ""
    if hasattr(value, "value"):   # Python enum instance
        return str(value.value)
    return str(value)
