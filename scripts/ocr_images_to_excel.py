"""
One-off utility: run the project's configured OCR provider (PaddleOCR,
same dual English/Cyrillic fusion engine used by the blade rocking system)
over every image in a folder and produce an Excel workbook with the image
thumbnail next to its detected text.

Usage (run from backend/ with its venv active):
    python ../scripts/ocr_images_to_excel.py <images_dir> <output_xlsx>
"""

from __future__ import annotations

import asyncio
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))

from app.ocr.paddle_provider import PaddleOCRProvider  # noqa: E402

THUMB_WIDTH = 240
ROW_HEIGHT_PER_PX = 0.75  # approx points per pixel for row height


async def main(images_dir: Path, output_path: Path) -> None:
    provider = PaddleOCRProvider()

    image_paths = sorted(
        p for p in images_dir.iterdir()
        if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")
    )
    if not image_paths:
        print(f"No images found in {images_dir}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Results"
    ws.append(["Image", "Filename", "Detected Text", "Confidence", "Provider", "Error"])
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 30

    for i, img_path in enumerate(image_paths, start=2):
        print(f"[{i - 1}/{len(image_paths)}] OCR {img_path.name} ...")
        image_bytes = img_path.read_bytes()
        result = await provider.extract_text(image_bytes)

        with PILImage.open(img_path) as pil_img:
            pil_img = pil_img.convert("RGB")
            w, h = pil_img.size
            scale = THUMB_WIDTH / w
            thumb_h = int(h * scale)
            pil_img = pil_img.resize((THUMB_WIDTH, thumb_h))
            thumb_path = output_path.parent / f"_thumb_{img_path.stem}.png"
            pil_img.save(thumb_path)

        xl_img = XLImage(str(thumb_path))
        xl_img.width = THUMB_WIDTH
        xl_img.height = thumb_h
        cell = f"A{i}"
        ws.add_image(xl_img, cell)
        ws.row_dimensions[i].height = max(thumb_h * ROW_HEIGHT_PER_PX, 15)

        ws.cell(row=i, column=2, value=img_path.name)
        ws.cell(row=i, column=3, value=result.raw_text)
        ws.cell(row=i, column=4, value=round(result.confidence, 3))
        ws.cell(row=i, column=5, value=result.provider)
        ws.cell(row=i, column=6, value=result.error)

    wb.save(output_path)

    # clean up temp thumbnails
    for p in output_path.parent.glob("_thumb_*.png"):
        p.unlink()

    print(f"Saved {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python ocr_images_to_excel.py <images_dir> <output_xlsx>")
        sys.exit(1)
    asyncio.run(main(Path(sys.argv[1]), Path(sys.argv[2])))
